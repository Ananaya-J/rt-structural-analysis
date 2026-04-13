#!/usr/bin/env python3
"""
01d_import_manual_annotations.py
================================
Convert manually curated domain annotations (CSV/XLSX) into dataset.json format
used by the training pipeline, with optional merge into an existing dataset.

Accepted input (one row per domain span):
  accession,sequence,domain,start,end[,pfam_id]

Typical usage:
  python 01d_import_manual_annotations.py \
      --input ./retroviral_annotations.xlsx \
      --sheet Sheet1 \
      --base_dataset ./training_data/dataset.json \
      --output ./training_data/dataset_augmented.json \
      --prefer manual
"""

import argparse
import csv
import json
from collections import defaultdict
from pathlib import Path


DOMAIN_NORMALIZATION = {
    "RVT_1": "RVT_1",
    "RVT1": "RVT_1",
    "Fingers": "RVT_1",
    "Palm": "RVT_1",
    "RVT_thumb": "RVT_thumb",
    "Thumb": "RVT_thumb",
    "RVT_connect": "RVT_connect",
    "Connection": "RVT_connect",
    "RNase_H": "RNase_H",
    "RNaseH": "RNase_H",
    "GIIM": "GIIM",
    "Maturase": "GIIM",
}


PFAM_BY_DOMAIN = {
    "RVT_1": "PF00078",
    "RVT_thumb": "PF06817",
    "RVT_connect": "PF06815",
    "RNase_H": "PF00075",
    "GIIM": "PF08388",
}


def normalize_domain(name):
    if name in DOMAIN_NORMALIZATION:
        return DOMAIN_NORMALIZATION[name]
    low = name.strip().lower()
    for key, value in DOMAIN_NORMALIZATION.items():
        if key.lower() == low:
            return value
    return None


def read_rows(input_path, sheet_name):
    path = Path(input_path)
    suffix = path.suffix.lower()
    if suffix in [".csv", ".tsv"]:
        delimiter = "\t" if suffix == ".tsv" else ","
        with open(path, newline="") as f:
            return list(csv.DictReader(f, delimiter=delimiter))

    if suffix in [".xlsx", ".xlsm"]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("XLSX input requires openpyxl: pip install openpyxl")

        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(x).strip() if x is not None else "" for x in rows[0]]
        out = []
        for row in rows[1:]:
            out.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
        return out

    raise ValueError(f"Unsupported input type: {suffix}. Use csv/tsv/xlsx")


def build_dataset_from_rows(
    rows,
    id_col,
    sequence_col,
    domain_col,
    start_col,
    end_col,
    pfam_col,
    base_sequences=None,
):
    grouped = defaultdict(list)
    sequences = {}
    skipped = 0
    unknown_domain_rows = 0
    inferred_sequence_rows = 0
    base_sequences = base_sequences or {}

    for row in rows:
        acc = str(row.get(id_col, "")).strip()
        seq = str(row.get(sequence_col, "")).strip().upper()
        domain_raw = str(row.get(domain_col, "")).strip()
        start = row.get(start_col, "")
        end = row.get(end_col, "")
        pfam_id = str(row.get(pfam_col, "")).strip() if pfam_col else ""

        if not acc or not domain_raw or start in ("", None) or end in ("", None):
            skipped += 1
            continue

        # For compact annotation tables, sequence may be omitted.
        # If a base dataset is provided, infer sequence by accession.
        if not seq and acc in base_sequences:
            seq = base_sequences[acc]
            inferred_sequence_rows += 1

        if not seq:
            skipped += 1
            continue

        norm_domain = normalize_domain(domain_raw)
        if norm_domain is None:
            unknown_domain_rows += 1
            skipped += 1
            continue

        try:
            start_i = int(float(start))
            end_i = int(float(end))
        except ValueError:
            skipped += 1
            continue

        if start_i > end_i or start_i < 1:
            skipped += 1
            continue

        prev = sequences.get(acc)
        if prev is None:
            sequences[acc] = seq
        elif prev != seq:
            # Inconsistent sequence for same accession, keep the longest
            sequences[acc] = seq if len(seq) > len(prev) else prev

        grouped[acc].append(
            {
                "domain_name": norm_domain,
                "pfam_id": pfam_id or PFAM_BY_DOMAIN.get(norm_domain, ""),
                "start": start_i,
                "end": end_i,
            }
        )

    dataset = []
    for acc, domains in grouped.items():
        seq = sequences[acc]
        labels = ["none"] * len(seq)

        # Stable and deterministic domain painting
        clipped_domains = []
        for d in sorted(domains, key=lambda x: (x["start"], x["end"])):
            start_i = max(1, d["start"])
            end_i = min(d["end"], len(seq))
            if start_i > len(seq) or start_i > end_i:
                continue
            clipped = dict(d)
            clipped["start"] = start_i
            clipped["end"] = end_i
            clipped_domains.append(clipped)

            for i in range(start_i - 1, end_i):
                if labels[i] == "none":
                    labels[i] = d["domain_name"]

        dataset.append(
            {
                "accession": acc,
                "sequence": seq,
                "labels": labels,
                "domains": clipped_domains,
                "source": "manual_import",
            }
        )

    stats = {
        "skipped_rows": skipped,
        "unknown_domain_rows": unknown_domain_rows,
        "inferred_sequence_rows": inferred_sequence_rows,
    }
    return dataset, stats


def merge_datasets(base, manual, prefer):
    by_acc = {d["accession"]: d for d in base}
    replaced = 0
    added = 0

    for item in manual:
        acc = item["accession"]
        if acc in by_acc:
            if prefer == "manual":
                by_acc[acc] = item
                replaced += 1
        else:
            by_acc[acc] = item
            added += 1

    return list(by_acc.values()), added, replaced


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--input", required=True, help="Manual annotation table (csv/tsv/xlsx)")
    p.add_argument("--sheet", default="", help="Excel sheet name (xlsx only)")
    p.add_argument("--base_dataset", default="", help="Existing dataset.json to merge into")
    p.add_argument("--output", required=True, help="Output dataset.json")
    p.add_argument("--prefer", choices=["manual", "base"], default="manual")
    p.add_argument("--id_col", default="accession")
    p.add_argument("--sequence_col", default="sequence")
    p.add_argument("--domain_col", default="domain")
    p.add_argument("--start_col", default="start")
    p.add_argument("--end_col", default="end")
    p.add_argument("--pfam_col", default="pfam_id")
    args = p.parse_args()

    rows = read_rows(args.input, args.sheet)
    base = []
    base_sequences = {}
    if args.base_dataset:
        with open(args.base_dataset) as f:
            base = json.load(f)
        base_sequences = {d["accession"]: d["sequence"] for d in base}

    manual_dataset, stats = build_dataset_from_rows(
        rows,
        id_col=args.id_col,
        sequence_col=args.sequence_col,
        domain_col=args.domain_col,
        start_col=args.start_col,
        end_col=args.end_col,
        pfam_col=args.pfam_col,
        base_sequences=base_sequences,
    )

    if args.base_dataset:
        merged, added, replaced = merge_datasets(base, manual_dataset, args.prefer)
        output_dataset = merged
        print(
            f"Merged dataset: base={len(base)}, manual={len(manual_dataset)}, "
            f"added={added}, replaced={replaced}, total={len(output_dataset)}"
        )
    else:
        output_dataset = manual_dataset
        print(f"Built manual dataset with {len(output_dataset)} proteins")

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(output_dataset, f, indent=2)

    print(f"Skipped rows: {stats['skipped_rows']}")
    print(f"Rows skipped for unknown domain names: {stats['unknown_domain_rows']}")
    print(f"Rows that inferred sequence from base dataset: {stats['inferred_sequence_rows']}")
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
